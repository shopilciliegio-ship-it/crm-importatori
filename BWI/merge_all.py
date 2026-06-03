"""
merge_all.py
============
Fonde tutti i file Excel nella cartella bestwine_output
in un unico file "bestwine_ALL.xlsx" con un foglio per ogni paese.
Poi puoi importare quel singolo file nel CRM.

Esegui: python merge_all.py
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_FOLDER  = "bestwine_output"
OUTPUT_FILE    = "bestwine_ALL.xlsx"

FIELDNAMES = [
    "CompanyName", "BrandName", "CompId",
    "Country", "City", "State", "StreetAddress", "PostalCode",
    "Website", "Type", "ProdType", "Employee", "Sales",
    "Company_Email", "Phone", "Founded", "RegistrationNumber",
    "Linkedin", "Facebook", "Instagram", "Twitter", "Youtube",
    "Contact_Name", "Contact_Title", "Contact_Email",
    "Contact_Phone", "Contact_Linkedin"
]

HEADER_FILL = PatternFill("solid", fgColor="4F1A2A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NO_FILL     = PatternFill(fill_type=None)
NO_FONT     = Font(color="000000", bold=False)

files = sorted([
    f for f in os.listdir(OUTPUT_FOLDER)
    if f.endswith(".xlsx") and "CORROTTO" not in f and f != OUTPUT_FILE
])

print(f"Trovati {len(files)} file da unire...\n")

wb_out = Workbook()
wb_out.remove(wb_out.active)  # rimuovi foglio vuoto default

total_rows = 0
skipped    = 0

for filename in files:
    path = os.path.join(OUTPUT_FOLDER, filename)
    # Nome del paese dal nome file (es. bestwine_Italy.xlsx → Italy)
    country_name = filename.replace("bestwine_", "").replace(".xlsx", "")
    # Excel: nome foglio max 31 caratteri
    sheet_name = country_name[:31]

    try:
        wb_in = load_workbook(path, read_only=True, data_only=True)
        ws_in = wb_in.active

        # Leggi tutte le righe (salta header)
        rows = list(ws_in.iter_rows(min_row=2, values_only=True))
        wb_in.close()

        if not rows:
            print(f"  ⬜ {filename} — vuoto, saltato")
            skipped += 1
            continue

        # Crea foglio di output
        ws_out = wb_out.create_sheet(title=sheet_name)

        # Header
        ws_out.append(FIELDNAMES)
        for cell in ws_out[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Dati (senza colori)
        import re
        def clean(val):
            s = str(val or "")
            return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', s)

        for row in rows:
            # Mappa le colonne del file sorgente alle colonne standard
            # (usa i primi 27 valori, padding con "" se mancano)
            clean_row = [clean(row[i] if i < len(row) else "") for i in range(len(FIELDNAMES))]
            ws_out.append(clean_row)

        total_rows += len(rows)
        print(f"  ✅ {sheet_name}: {len(rows)} righe")

    except Exception as e:
        print(f"  ⚠️  {filename}: {e}")
        skipped += 1

# Salva
output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
wb_out.save(output_path)

print(f"""
{'='*50}
✅ FATTO!
   File: {output_path}
   Fogli (paesi): {len(wb_out.sheetnames)}
   Righe totali:  {total_rows:,}
   File saltati:  {skipped}
{'='*50}
Ora importa bestwine_ALL.xlsx nel CRM con un solo drag & drop.
""")
