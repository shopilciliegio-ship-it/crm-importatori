"""
BestWineImporters - Scraper Aziende e Contatti
================================================
Il login avviene in automatico tramite bwi_auto_login.py.
Non serve più copiare cookie o token a mano.

Uso interattivo:   python bestwine_scraper.py
Uso automatico:    python bestwine_scraper.py --auto 2
  (modalità 2 = nuove+aggiornate, usato dal cron settimanale)

I file Excel vengono salvati in "bestwine_output/",
uno per ogni nazione (es. bestwine_Italy.xlsx).
"""

import sys
import requests
import time
import os
import threading
from datetime import datetime, timezone
from requests.adapters import HTTPAdapter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Auto-login — sovrascrive i valori qui sotto all'avvio di main()
try:
    from bwi_auto_login import do_login as _bwi_login
    _AUTO_LOGIN = True
except ImportError:
    _AUTO_LOGIN = False

# Valori di fallback (usati solo se auto-login non disponibile)
COOKIE     = ""
BDN_ACCESS = ""
BDN_ID     = "13229"
BDN_NAME   = "info@sienawine.it"

# ============================================================
# FILTRI DI RICERCA
# Lascia le liste vuote [] per scaricare tutto
# Esempio per un paese solo: "FilterCountry": ["Italy"]
# ============================================================

FILTRI = {
    "FilterKeyword": [],
    "FilterCategories": [],
    "FilterSubCategories": [],
    "FilterTypes": [],
    "FilterOrigin": [],
    "FilterCountry": [],
    "FilterContinent": [],
    "FilterEmployee": [],
    "FilterSales": []
}

# Cartella di output
OUTPUT_FOLDER  = "bestwine_output"
TIMESTAMP_FILE = "ultimo_aggiornamento.txt"  # salva data ultimo run

# MODALITÀ AGGIORNAMENTO
# False = salta le aziende già scaricate (veloce, solo nuove)
# True  = riscarica tutto e aggiorna i dati esistenti (lento, ma aggiorna)
FORCE_UPDATE = False

# MODALITÀ DI SCARICO
# "tutto"      = scarica tutto il database (comportamento normale)
# "nuove"      = scarica solo le aziende aggiunte di recente  (/createdcomp)
# "aggiornate" = scarica solo le aziende modificate di recente (/updatedcomp)
MODALITA = "tutto"

# Usato solo con MODALITA="nuove" o "aggiornate"
# Quante aziende recenti prendere (le ultime N, ordinate per data desc)
# 0 = tutte quelle disponibili nell'endpoint
ULTIME_N = 0

# FILTRO PAESE (solo MODALITA="tutto")
# "" = tutti i paesi (normale)
# "France" = riscarica solo la Francia (utile per file corrotti)
SOLO_PAESE = "" 

# Pausa in secondi tra una azienda e l'altra
PAUSA_PROFILO = 0.3
PAUSA_LEADS   = 0.3

# Colore righe nuove (verde chiaro)
NUOVO_FILL = PatternFill("solid", fgColor="C6EFCE")
NUOVO_FONT = Font(color="276221")

# ============================================================

BASE_URL = "https://api.bestwineimporters.com/api/v1"

def load_last_run():
    """Legge la data dell'ultimo run da file. Ritorna None se primo avvio."""
    if not os.path.exists(TIMESTAMP_FILE):
        return None
    try:
        with open(TIMESTAMP_FILE, "r") as f:
            ts = f.read().strip()
            return datetime.fromisoformat(ts)
    except Exception:
        return None

def save_last_run():
    """Salva la data/ora attuale come ultimo run."""
    with open(TIMESTAMP_FILE, "w") as f:
        f.write(datetime.now(timezone.utc).isoformat())

def parse_date(date_str):
    """Converte stringa ISO in datetime aware."""
    if not date_str:
        return None
    try:
        # Gestisce formati con e senza Z finale
        ds = date_str.replace("Z", "+00:00")
        return datetime.fromisoformat(ds)
    except Exception:
        return None

def get_recent_companies(endpoint):
    """
    Scarica aziende nuove o aggiornate usando customSearch con sortBy.
    endpoint: "createdcomp" → sortBy Created DESC
              "updatedcomp" → sortBy Updated DESC
    Scarica a blocchi da 50 finché non trova aziende più vecchie del cutoff.
    """
    sort_field = "Created" if endpoint == "createdcomp" else "Updated"
    last_run   = load_last_run()

    # Se è il primo run, limita a 1000 aziende per evitare di scaricare tutto
    max_items  = 99999 if last_run else 1000

    all_items  = []
    start      = 0
    batch      = 50
    stop_early = False

    print(f"   Ordinamento: {sort_field} DESC", end="")
    if last_run:
        print(f" | Cutoff: {last_run.strftime('%Y-%m-%d %H:%M')} UTC")
    else:
        print(f" | Primo run — massimo {max_items} aziende")

    while start < max_items:
        payload = {
            "nocache": False,
            "startAt": start,
            "endAt":   start + batch,
            "sortBy":  sort_field,
            "sortDir": "DESC",
            "filters": FILTRI,
            "id": None, "name": None, "date": None
        }
        r, err = call_api(f"{BASE_URL}/customSearch/", payload, timeout_sec=15)
        if err:
            print(f"\n❌ Errore customSearch: {err}")
            break
        if r.status_code in (401, 403):
            print("\n❌ AUTENTICAZIONE SCADUTA — aggiorna COOKIE e rilancia.")
            raise SystemExit(1)
        try:
            data     = r.json()
            total    = data.get("countComp",[{}])[0].get("TotalComp", 0)
            companies = data.get("resListComp", [])
        except Exception as e:
            print(f"\n❌ JSON non valido: {e}")
            break

        if not companies:
            break

        for comp in companies:
            # Usa la data di Created o Updated dal campo comp
            date_str = comp.get("Created") or comp.get("Updated") or ""
            comp_date = parse_date(date_str)

            # Se abbiamo un cutoff e la data è più vecchia → stop
            if last_run and comp_date and comp_date <= last_run:
                stop_early = True
                break

            # Normalizza il formato (customSearch ha campi diversi da createdcomp)
            item = {
                "CompId":       comp.get("CompId"),
                "CompanyName":  comp.get("CompanyName",""),
                "Country":      comp.get("Country",""),
                "City":         comp.get("City",""),
                "State":        comp.get("State",""),
                "Type":         comp.get("Type",""),
                "ProdType":     comp.get("ProdType",""),
                "Website":      comp.get("Website",""),
                "Employee":     comp.get("Employee",""),
                "Sales":        comp.get("Sales",""),
                "Date":         date_str,
            }
            all_items.append(item)

        print(f"   {len(all_items)}/{total}...", end="\r")

        if stop_early or start + batch >= total or start + batch >= max_items:
            break
        start += batch

    print()  # newline dopo i progressi
    return all_items

def make_session():
    """Crea una session HTTP fresca — evita connessioni appese."""
    s = requests.Session()
    adapter = HTTPAdapter(max_retries=0)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s

def call_api(url, payload, timeout_sec=10):
    """Chiama l'API con hard timeout via thread — non si blocca mai."""
    result = [None]
    error  = [None]

    def _call():
        try:
            with make_session() as sess:
                if payload is None:
                    r = sess.get(url, headers=HEADERS, timeout=(5, timeout_sec))
                else:
                    r = sess.post(url, json=payload, headers=HEADERS, timeout=(5, timeout_sec))
                result[0] = r
        except Exception as e:
            error[0] = e

    t = threading.Thread(target=_call, daemon=True)
    t.start()
    t.join(timeout=timeout_sec + 3)  # hard timeout = timeout HTTP + 3s margine

    if t.is_alive():
        # Thread ancora appeso — restituisce None (verrà trattato come errore)
        return None, TimeoutError(f"Hard timeout {timeout_sec+3}s superato")
    if error[0]:
        return None, error[0]
    return result[0], None

HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Content-Type": "application/json",
    "Cookie": COOKIE,
    "Bdn-Access": BDN_ACCESS,
    "Bdn-Id": BDN_ID,
    "Bdn-Name": BDN_NAME,
    "Origin": "https://app.bestwineimporters.com",
    "Referer": "https://app.bestwineimporters.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Connection": "close"  # chiude la connessione dopo ogni richiesta, evita hang
}

FIELDNAMES = [
    "CompanyName", "BrandName", "CompId",
    "Country", "City", "State", "StreetAddress", "PostalCode",
    "Website", "Type", "ProdType", "Employee", "Sales",
    "Company_Email", "Phone", "Founded", "RegistrationNumber",
    "Linkedin", "Facebook", "Instagram", "Twitter", "Youtube",
    "Contact_Name", "Contact_Title", "Contact_Email",
    "Contact_Phone", "Contact_Linkedin"
]

COMPID_COL_INDEX = FIELDNAMES.index("CompId")


def get_batch(start, end):
    """Scarica un blocco di aziende dalla lista — con retry automatico."""
    payload = {
        "nocache": False,
        "startAt": start,
        "endAt": end,
        "sortBy": "",
        "sortDir": "",
        "filters": FILTRI,
        "id": None,
        "name": None,
        "date": None
    }
    for attempt in range(5):
        r, err = call_api(f"{BASE_URL}/customSearch/", payload, timeout_sec=10)
        if err:
            print(f"  ⏱  customSearch tentativo {attempt+1}/5: {err}")
            time.sleep(8 * (attempt + 1))
            continue
        if r.status_code in (401, 403):
            print("\n❌ AUTENTICAZIONE SCADUTA — aggiorna COOKIE e rilancia.")
            raise SystemExit(1)
        if r.status_code == 429:
            print(f"\n⏳ Rate limit — aspetto {60*(attempt+1)}s...")
            time.sleep(60 * (attempt + 1))
            continue
        try:
            return r.json()
        except Exception as e:
            print(f"  ⚠️  JSON non valido: {e}")
            time.sleep(5)
    print("\n❌ customSearch fallita dopo 5 tentativi.")
    raise SystemExit(1)


def get_profile(comp_id):
    r, err = call_api(f"{BASE_URL}/profileinfo/", {"compID": comp_id}, timeout_sec=8)
    if err:
        print(f"  ⏱  Profilo timeout/errore: {err}")
        return {}
    if r.status_code in (401, 403):
        # 401 su singola azienda = profilo ristretto o legacy → skip
        print(f"  ⚠️  Profilo non accessibile (ID: {comp_id}) — salto")
        return {}
    try:
        return r.json()
    except Exception:
        return {}


def get_leads(comp_id):
    r, err = call_api(f"{BASE_URL}/leads", {"compID": comp_id}, timeout_sec=8)
    if err:
        print(f"  ⏱  Leads timeout/errore: {err}")
        return []
    if r.status_code in (401, 403):
        print("\n❌ AUTENTICAZIONE SCADUTA — aggiorna il token e rilancia.")
        raise SystemExit(1)
    try:
        data = r.json()
        if isinstance(data, list): return data
        if isinstance(data, dict): return data.get("leads", data.get("resListLeads", []))
    except Exception:
        pass
    return []


def extract_phone(phones):
    if not phones or not isinstance(phones, dict):
        return ""
    for key in ["fixed_line", "fixed"]:
        nums = phones.get(key, [])
        if nums and isinstance(nums, list):
            return nums[0]
    return ""


def clean_name(name):
    return "".join(c for c in name if c.isalnum() or c in (" ", "-", "_")).strip()


def get_filepath(country):
    return os.path.join(OUTPUT_FOLDER, f"bestwine_{clean_name(country)}.xlsx")


def setup_sheet(ws):
    ws.append(FIELDNAMES)
    header_fill = PatternFill("solid", fgColor="4F1A2A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    col_widths = [35, 25, 10, 15, 15, 20, 25, 12, 30, 25, 25, 12,
                  15, 28, 18, 10, 20, 35, 35, 35, 15, 35, 28, 25, 30, 18, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width


def load_existing_ids():
    """Legge tutti i CompId già salvati nei file xlsx esistenti."""
    already_done = set()
    if not os.path.exists(OUTPUT_FOLDER):
        return already_done
    files = [f for f in os.listdir(OUTPUT_FOLDER)
              if f.endswith(".xlsx")
              and "_OLD" not in f        # ignora backup rinominati
              and "CORROTTO" not in f    # ignora file corrotti
              and f != "bestwine_ALL.xlsx"]
    if not files:
        return already_done
    print(f"\n🔎 Trovati {len(files)} file esistenti — carico CompId già elaborati...")
    for filename in files:
        try:
            wb = load_workbook(os.path.join(OUTPUT_FOLDER, filename), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                cid = row[COMPID_COL_INDEX]
                if cid:
                    already_done.add(str(cid))
            wb.close()
            print(f"  ✅ {filename}")
        except Exception as e:
            print(f"  ⚠️  Errore lettura {filename}: {e}")
            print(f"     → File probabilmente corrotto. Cancellalo o rinominalo e rilancia.")
            if "zip" in str(e).lower() or "not a zip" in str(e).lower():
                print(f"  🗑️  Il file sembra corrotto. Eliminalo e rilancia per riscaricarlo.")
    print(f"  📋 Totale già elaborate: {len(already_done)} — verranno saltate.\n")
    return already_done


def get_or_create_workbook(workbooks, country):
    """Apre il file xlsx esistente oppure ne crea uno nuovo."""
    if country not in workbooks:
        filepath = get_filepath(country)
        if os.path.exists(filepath):
            try:
                wb = load_workbook(filepath)
                ws = wb.active
                workbooks[country] = {"wb": wb, "ws": ws, "is_update": True}
            except Exception as e:
                print(f"  ⚠️  File corrotto per {country}: {e}")
                print(f"  🔄  Creo un nuovo file — i dati precedenti di {country} verranno riscritti.")
                # Rinomina il file corrotto invece di eliminarlo
                import shutil
                backup = filepath.replace(".xlsx", "_CORROTTO.xlsx")
                shutil.move(filepath, backup)
                print(f"  💾  File corrotto rinominato in: {os.path.basename(backup)}")
                wb = Workbook()
                ws = wb.active
                ws.title = country[:31]
                setup_sheet(ws)
                workbooks[country] = {"wb": wb, "ws": ws, "is_update": False}
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = country[:31]
            setup_sheet(ws)
            workbooks[country] = {"wb": wb, "ws": ws, "is_update": False}
    return workbooks[country]


def write_row(ws, data, highlight=False):
    import re

    def clean(val):
        s = str(val or "")
        # Rimuove caratteri di controllo illegali per Excel (U+0000–U+001F escluso tab/newline)
        return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', s)

    ws.append([clean(data.get(f, "")) for f in FIELDNAMES])
    if highlight:
        for cell in ws[ws.max_row]:
            cell.fill = NUOVO_FILL
            cell.font = NUOVO_FONT


def save_all(workbooks):
    for country, obj in workbooks.items():
        obj["wb"].save(get_filepath(country))


# Contatore errori consecutivi — se troppi, pausa lunga
_errori_consecutivi = 0

def process_company(comp, workbooks, is_update_run, counter):
    """Elabora una singola azienda: profilo + leads + scrittura Excel."""
    global _errori_consecutivi
    comp_id   = comp.get("CompId")
    comp_name = comp.get("CompanyName", "N/A")
    print(f"  [{counter}] {comp_name} (ID: {comp_id})")

    profile = get_profile(comp_id)
    time.sleep(PAUSA_PROFILO)

    social   = profile.get("social", {})
    phones   = profile.get("phones", {})
    datadnb  = profile.get("datadnb", {})
    location = profile.get("location", {})

    country = location.get("Country", comp.get("Country", "Unknown")) or "Unknown"

    base = {
        "CompanyName":        comp_name,
        "BrandName":          profile.get("brandname", ""),
        "CompId":             comp_id,
        "Country":            country,
        "City":               location.get("City", comp.get("City", "")),
        "State":              location.get("State", comp.get("State", "")),
        "StreetAddress":      location.get("Address", comp.get("StreetAddress", "")),
        "PostalCode":         location.get("Postal", comp.get("PostalCode", "")),
        "Website":            profile.get("website", comp.get("Website", "")),
        "Type":               comp.get("Type", ""),
        "ProdType":           comp.get("ProdType", ""),
        "Employee":           datadnb.get("employees", comp.get("Employee", "")),
        "Sales":              datadnb.get("sales", comp.get("Sales", "")),
        "Company_Email":      profile.get("email", ""),
        "Phone":              extract_phone(phones),
        "Founded":            datadnb.get("Founded", ""),
        "RegistrationNumber": datadnb.get("registrationNumber", ""),
        "Linkedin":           social.get("Linkedin", comp.get("Linkedin", "")),
        "Facebook":           social.get("Facebook", comp.get("Facebook", "")),
        "Instagram":          social.get("Instagram", comp.get("Instagram", "")),
        "Twitter":            social.get("Twitter", ""),
        "Youtube":            social.get("Youtube", ""),
    }

    leads = get_leads(comp_id)
    time.sleep(PAUSA_LEADS)

    # Reset contatore errori dopo elaborazione riuscita
    _errori_consecutivi = 0

    obj = get_or_create_workbook(workbooks, country)
    ws  = obj["ws"]
    highlight = obj["is_update"]

    if leads:
        for lead in leads:
            row = base.copy()
            row["Contact_Name"]     = lead.get("FullName", "")
            row["Contact_Title"]    = lead.get("Position", "")
            row["Contact_Email"]    = ""   # FASE 1: nessuno sblocco email
            row["Contact_Phone"]    = lead.get("Phone", "")
            row["Contact_Linkedin"] = lead.get("LinkedIn", "")
            write_row(ws, row, highlight=highlight)
    else:
        write_row(ws, base, highlight=highlight)


def main():
    global MODALITA, SOLO_PAESE, ULTIME_N, _errori_consecutivi, HEADERS

    # ── AUTO-LOGIN ────────────────────────────────────────────────────
    if _AUTO_LOGIN:
        print("🔑 Auto-login BWI...", flush=True)
        try:
            _sess, _h = _bwi_login()
            HEADERS = _h
            print("   Login OK")
        except Exception as e:
            print(f"   Login FALLITO: {e}")
            raise SystemExit(1)
    else:
        print("⚠  bwi_auto_login non disponibile — usa token statici nel file.")

    # ── MODALITÀ: da riga di comando o interattiva ────────────────────
    # --auto N  salta il menu (N = 1-4, usato dal cron GitHub Actions)
    auto_mode = None
    if "--auto" in sys.argv:
        try:
            auto_mode = sys.argv[sys.argv.index("--auto") + 1]
        except IndexError:
            auto_mode = "2"

    print("=" * 60)
    print("  BestWineImporters - Scraper")
    print("=" * 60)
    print()

    if auto_mode:
        scelta = auto_mode
        print(f"  Modalità automatica: [{scelta}]")
    else:
        print("  Scegli modalità:")
        print("  [1] TUTTO              — scarica tutto il database")
        print("  [2] NUOVE + AGGIORNATE — nuove E modificate (uso settimanale)")
        print("  [3] Solo NUOVE         — solo /createdcomp")
        print("  [4] Solo AGGIORNATE    — solo /updatedcomp")
        print(f"  [5] Default dal file ({MODALITA.upper()})")
        print()
        scelta = input("  Scelta [1-5, invio = default]: ").strip()

    if scelta == "1":
        MODALITA = "tutto"
    elif scelta == "2":
        MODALITA = "nuove+aggiornate"
    elif scelta == "3":
        MODALITA = "nuove"
        if not auto_mode:
            n = input("  Quante? [invio = tutte]: ").strip()
            if n.isdigit(): ULTIME_N = int(n)
    elif scelta == "4":
        MODALITA = "aggiornate"
        if not auto_mode:
            n = input("  Quante? [invio = tutte]: ").strip()
            if n.isdigit(): ULTIME_N = int(n)
    # else: usa default

    if MODALITA == "tutto" and not auto_mode:
        paese = input("  Filtra per paese? [invio = tutti, es: France]: ").strip()
        SOLO_PAESE = paese
    else:
        SOLO_PAESE = ""

    print()
    label_map = {"tutto":"TUTTO","nuove":"Solo NUOVE",
                 "aggiornate":"Solo AGGIORNATE","nuove+aggiornate":"NUOVE + AGGIORNATE"}
    print(f"  ✅ Modalità: {label_map.get(MODALITA, MODALITA.upper())}", end="")
    if SOLO_PAESE: print(f"  |  Paese: {SOLO_PAESE}", end="")
    print()
    print("=" * 60)
    print()

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # ── MODALITÀ NUOVE / AGGIORNATE / NUOVE+AGGIORNATE ─────────────
    if MODALITA in ("nuove", "aggiornate", "nuove+aggiornate"):

        # Determina quali endpoint chiamare
        endpoints = []
        if MODALITA == "nuove":
            endpoints = [("createdcomp", "NUOVE")]
        elif MODALITA == "aggiornate":
            endpoints = [("updatedcomp", "AGGIORNATE")]
        else:  # nuove+aggiornate
            endpoints = [("createdcomp", "NUOVE"), ("updatedcomp", "AGGIORNATE")]

        already_done = load_existing_ids()  # per skippare nuove già presenti
        workbooks    = {}
        elaborated   = 0
        seen_ids     = set()  # evita duplicati se un ID compare in entrambi

        for endpoint, label in endpoints:
            print(f"🔍 Recupero aziende {label} da /{endpoint}...")
            items = get_recent_companies(endpoint)
            if not items:
                print(f"   Nessuna azienda {label.lower()} trovata.")
                continue

            if ULTIME_N > 0:
                items = items[:ULTIME_N]

            print(f"✅ {len(items)} aziende {label} da elaborare")
            if items:
                dates = [i.get('Date','?')[:10] for i in items if i.get('Date')]
                if dates:
                    print(f"   Da: {min(dates)}  A: {max(dates)}")
            print()

            for item in items:
                comp_id = str(item.get("CompId", ""))
                if not comp_id: continue

                # Salta duplicati tra i due endpoint
                if comp_id in seen_ids:
                    continue
                seen_ids.add(comp_id)

                # Per le NUOVE: salta se già in archivio
                # Per le AGGIORNATE: processa sempre (dati cambiati)
                if label == "NUOVE" and comp_id in already_done:
                    print(f"  ⏭  {item.get('CompanyName','')} — già presente")
                    continue

                comp = {
                    "CompId":      item.get("CompId"),
                    "CompanyName": item.get("CompanyName", ""),
                    "Country":     item.get("Country", ""),
                    "City":        item.get("City", ""),
                    "State":       "",
                    "Type":        item.get("CompanyType", ""),
                    "ProdType":    item.get("Categories", ""),
                    "Website":     item.get("Website", ""),
                    "Employee":    "",
                    "Sales":       "",
                }

                elaborated += 1
                process_company(comp, workbooks, True, elaborated)

                if _errori_consecutivi > 3:
                    print(f"\n⚠️  Troppi errori — pausa 60s...")
                    time.sleep(60)
                    _errori_consecutivi = 0

        print(f"\n💾 Salvataggio finale...")
        save_all(workbooks)
        save_last_run()  # aggiorna timestamp solo se tutto ok

        # Scrive manifest con i CompId elaborati in questo run
        # → usato da bwi_sync_crm.py per un sync incrementale (non tutto l'archivio)
        manifest_path = os.path.join(OUTPUT_FOLDER, "last_sync_ids.json")
        with open(manifest_path, "w") as f:
            import json as _json
            _json.dump(list(seen_ids), f)
        print(f"   Manifest: {len(seen_ids)} ID scritti in last_sync_ids.json")

        print(f"\n🎉 FATTO! {elaborated} aziende elaborate.")
        print(f"   ⏰ Prossimo run: verranno saltate le aziende precedenti a ora.")
        return

    # ── MODALITÀ TUTTO ───────────────────────────────────────────────
    if SOLO_PAESE:
        paese_file = os.path.join(OUTPUT_FOLDER, f"bestwine_{clean_name(SOLO_PAESE)}.xlsx")
        if os.path.exists(paese_file):
            backup = paese_file.replace(".xlsx", "_OLD.xlsx")
            os.rename(paese_file, backup)
            print(f"  📦 Backup: {os.path.basename(backup)}")
            print(f"  🔄 Riscarico: {SOLO_PAESE}")
            print()

    already_done = load_existing_ids()
    SKIP_IDS = set()
    already_done.update(SKIP_IDS)
    is_update_run = len(already_done) > 0

    print("🟢 Modalità AGGIORNAMENTO — nuove righe in verde." if is_update_run else "🆕 Prima esecuzione.")

    print("\n🔍 Recupero totale aziende...")
    first = get_batch(0, 1)
    total = first.get("countComp", [{}])[0].get("TotalComp", 0)
    print(f"✅ Totale: {total:,} aziende")

    if total == 0:
        print("❌ Nessuna azienda trovata.")
        return

    print(f"\n📋 Avvio scarico (blocchi da 50)...\n")

    workbooks  = {}
    batch      = 50
    elaborated = 0

    for start in range(0, total, batch):
        end  = min(start + batch, total)
        page = get_batch(start, end)
        companies = page.get("resListComp", [])

        for comp in companies:
            comp_id      = str(comp.get("CompId"))
            comp_country = comp.get("Country", "")

            if SOLO_PAESE:
                if comp_country.lower() != SOLO_PAESE.lower():
                    continue
            else:
                if comp_id in already_done and not FORCE_UPDATE:
                    continue

            elaborated += 1
            process_company(comp, workbooks, is_update_run, elaborated)

            if _errori_consecutivi > 3:
                print(f"\n⚠️  Errori consecutivi — pausa 60s...")
                time.sleep(60)
                _errori_consecutivi = 0

            if elaborated % 100 == 0:
                save_all(workbooks)
                print(f"  💾 {elaborated} elaborate")

    print(f"\n💾 Salvataggio finale...")
    save_all(workbooks)
    print(f"\n🎉 FATTO! {elaborated} aziende elaborate.")
    if is_update_run:
        print("   Nuove righe in VERDE.")


if __name__ == "__main__":
    main()
