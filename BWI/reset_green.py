"""
reset_green.py
==============
Rimuove l'evidenziazione verde da tutte le righe dei file Excel
nella cartella bestwine_output (le righe "nuove" diventano normali).
Esegui: python reset_green.py
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

OUTPUT_FOLDER = "bestwine_output"
NO_FILL  = PatternFill(fill_type=None)
NO_FONT  = Font(color="000000", bold=False)

# Colori da rimuovere (verde = nuove, azzurro = recuperate da recovery)
COLORS_TO_REMOVE = {"C6EFCE", "BDD7EE"}  # aggiungi altri se necessario

files = [f for f in os.listdir(OUTPUT_FOLDER) if f.endswith(".xlsx")]
print(f"Trovati {len(files)} file Excel...")

for filename in files:
    path = os.path.join(OUTPUT_FOLDER, filename)
    try:
        wb = load_workbook(path)
        ws = wb.active
        modified = False
        # Dalla riga 2 in poi (salta header)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.fill and cell.fill.fill_type not in (None, "none"):
                    # Rimuove verde (nuove) e azzurro (recuperate) ma non altri colori
                    fg = cell.fill.fgColor.rgb if cell.fill.fgColor else ""
                    if fg in COLORS_TO_REMOVE or fg.endswith(tuple(COLORS_TO_REMOVE)):
                        cell.fill = NO_FILL
                        cell.font = NO_FONT
                        modified = True
        if modified:
            wb.save(path)
            print(f"  ✅ {filename} — verde rimosso")
        else:
            print(f"  ⬜ {filename} — già pulito")
    except Exception as e:
        print(f"  ⚠️  {filename}: {e}")

print("\nFatto!")
