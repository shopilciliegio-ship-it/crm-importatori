# -*- coding: utf-8 -*-
"""
bwi_sync_crm.py
===============
Legge bestwine_ALL.xlsx e aggiorna data/crm.json con i dati BWI.

Logica di merge (chiave: CompId):
  - Nuovo CompId  → aggiunge contatto, bwiStatus='new'
  - CompId esiste, dati cambiati → aggiorna campi, bwiStatus='updated', log
  - CompId esiste, dati identici → skip (tocca solo bwiSyncedAt)
  - brevoEvents non viene mai toccato dal sync

Mapping XLSX → CRM:
  CompanyName → company        Company_Email → email
  BrandName   → brandName      Phone         → phone
  Country     → country        Website       → website
  City        → city           Employee      → employees
  State       → region         Sales         → sales
  Type        → type           Linkedin      → linkedin
  ProdType    → prodType       Contact_*     → contacts[]

Alla fine invia digest a luca@ilciliegio.com via Brevo.
"""

import json, os, sys, re, time, requests
from datetime import datetime, timezone
from openpyxl import load_workbook

# ── PATH ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.dirname(SCRIPT_DIR)
CRM_FILE   = os.path.join(REPO_ROOT, "data", "contatti.json")
ALL_XLSX   = os.path.join(SCRIPT_DIR, "bestwine_output", "bestwine_ALL.xlsx")

# ── BREVO ─────────────────────────────────────────────────────────────
BREVO_API_KEY   = os.environ.get("BREVO_API_KEY", "")
DIGEST_TO_EMAIL = "luca@ilciliegio.com"
DIGEST_TO_NAME  = "Luca"
SENDER_EMAIL    = "luca@sienawine.it"
SENDER_NAME     = "Siena Wine CRM"

# ── CAMPI DA CONFRONTARE PER RILEVARE UN "AGGIORNAMENTO" ──────────────
COMPARE_FIELDS = ["company", "email", "phone", "website", "country", "city",
                  "employees", "sales", "brandName", "type", "prodType"]

# ─────────────────────────────────────────────────────────────────────

def _s(v) -> str:
    """Pulisce un valore di cella in stringa leggibile."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan", "-") else s


def load_xlsx_companies(xlsx_path: str) -> dict:
    """
    Legge bestwine_ALL.xlsx (un foglio per paese).
    Restituisce {CompId: {...}} con tutti i campi e contacts[].
    Più righe per stesso CompId = più contatti/lead per l'azienda.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    companies: dict = {}

    FIELDNAMES = [
        "CompanyName", "BrandName", "CompId",
        "Country", "City", "State", "StreetAddress", "PostalCode",
        "Website", "Type", "ProdType", "Employee", "Sales",
        "Company_Email", "Phone", "Founded", "RegistrationNumber",
        "Linkedin", "Facebook", "Instagram", "Twitter", "Youtube",
        "Contact_Name", "Contact_Title", "Contact_Email",
        "Contact_Phone", "Contact_Linkedin"
    ]

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[2]:  # CompId in colonna 3 (index 2)
                continue
            r = {FIELDNAMES[i]: _s(row[i] if i < len(row) else "") for i in range(len(FIELDNAMES))}
            cid = str(r["CompId"])

            contact_entry = {
                "name":    r["Contact_Name"],
                "title":   r["Contact_Title"],
                "email":   r["Contact_Email"],
                "phone":   r["Contact_Phone"],
                "linkedin":r["Contact_Linkedin"],
            }

            if cid not in companies:
                companies[cid] = {
                    "bwiCompId": cid,
                    "company":   r["CompanyName"],
                    "brandName": r["BrandName"],
                    "country":   r["Country"],
                    "city":      r["City"],
                    "region":    r["State"],
                    "website":   r["Website"],
                    "email":     r["Company_Email"],
                    "phone":     r["Phone"],
                    "employees": r["Employee"],
                    "sales":     r["Sales"],
                    "type":      r["Type"],
                    "prodType":  r["ProdType"],
                    "linkedin":  r["Linkedin"],
                    "contacts":  [],
                }

            # Aggiungi contatto solo se ha almeno nome o email
            if contact_entry["name"] or contact_entry["email"]:
                # Evita duplicati
                existing = companies[cid]["contacts"]
                if not any(ec["email"] == contact_entry["email"] and
                           ec["name"]  == contact_entry["name"]
                           for ec in existing):
                    existing.append(contact_entry)

    wb.close()
    return companies


def load_crm(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_crm(path: str, data: dict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


def _uid() -> str:
    import random, string
    ts = str(int(time.time() * 1000))
    rnd = "".join(random.choices(string.ascii_lowercase + string.digits, k=5))
    return f"c{ts}{rnd}"


def _contacts_changed(old_contacts: list, new_contacts: list) -> bool:
    """True se la lista contatti è diversa (confronto per email)."""
    old_emails = {(c.get("email",""), c.get("name","")) for c in old_contacts}
    new_emails = {(c.get("email",""), c.get("name","")) for c in new_contacts}
    return old_emails != new_emails


def sync(crm_data: dict, xlsx_companies: dict, now_ms: int) -> dict:
    """
    Merges xlsx_companies into crm_data.contacts.
    Returns stats dict.
    """
    contacts = crm_data.setdefault("contacts", [])

    # Indice per bwiCompId
    by_bwi: dict = {}
    for c in contacts:
        cid = c.get("bwiCompId")
        if cid:
            by_bwi[str(cid)] = c

    # Indice per nome azienda (migrazione vecchio import senza bwiCompId)
    # Vecchio import: company name era in campo "country"
    by_name_migration: dict = {}
    for c in contacts:
        if not c.get("bwiCompId"):
            key = (c.get("company","") or c.get("country","")).strip().lower()
            if key:
                by_name_migration[key] = c

    stats = {
        "new": 0, "updated": 0, "skipped": 0,
        "new_list": [], "updated_list": []
    }

    for cid, xlsx_c in xlsx_companies.items():
        existing = by_bwi.get(cid)

        # Fallback migration match per nome azienda
        if not existing:
            key = xlsx_c["company"].strip().lower()
            existing = by_name_migration.get(key)

        if existing:
            # Aggiorna bwiCompId se mancava (migration)
            existing["bwiCompId"] = cid

            # Controlla se qualcosa è cambiato
            changed_fields = []
            for field in COMPARE_FIELDS:
                old_val = str(existing.get(field) or "").strip()
                new_val = str(xlsx_c.get(field) or "").strip()
                if old_val != new_val and new_val:
                    changed_fields.append(field)

            contacts_changed = _contacts_changed(
                existing.get("contacts", []),
                xlsx_c.get("contacts", [])
            )

            if changed_fields or contacts_changed:
                # Aggiorna i campi da XLSX (non tocca status/brevoEvents/log)
                for field in COMPARE_FIELDS:
                    if xlsx_c.get(field):
                        existing[field] = xlsx_c[field]
                if contacts_changed and xlsx_c.get("contacts"):
                    existing["contacts"] = xlsx_c["contacts"]

                # Aggiorna contactName/contactEmail dal primo contatto
                _update_primary_contact(existing)

                existing["bwiStatus"]    = "updated"
                existing["bwiUpdatedAt"] = now_ms
                existing["updatedAt"]    = now_ms
                existing.setdefault("log", []).append({
                    "ts":  now_ms,
                    "msg": f"BWI sync: aggiornato ({', '.join(changed_fields or ['contatti'])})"
                })
                stats["updated"] += 1
                stats["updated_list"].append(existing["company"])
            else:
                existing["bwiSyncedAt"] = now_ms
                stats["skipped"] += 1

            # Rimuovi flag 'new' se il contatto ha ricevuto email
            if existing.get("bwiStatus") == "new" and existing.get("brevoEvents"):
                existing["bwiStatus"] = None
        else:
            # Nuovo contatto
            new_c = {
                "id":            _uid(),
                "bwiCompId":     cid,
                "company":       xlsx_c["company"],
                "brandName":     xlsx_c.get("brandName",""),
                "country":       xlsx_c.get("country",""),
                "city":          xlsx_c.get("city",""),
                "region":        xlsx_c.get("region",""),
                "website":       xlsx_c.get("website",""),
                "email":         xlsx_c.get("email",""),
                "phone":         xlsx_c.get("phone",""),
                "employees":     xlsx_c.get("employees",""),
                "sales":         xlsx_c.get("sales",""),
                "type":          xlsx_c.get("type",""),
                "prodType":      xlsx_c.get("prodType",""),
                "linkedin":      xlsx_c.get("linkedin",""),
                "contacts":      xlsx_c.get("contacts",[]),
                "status":        "new",
                "bwiStatus":     "new",
                "bwiImportedAt": now_ms,
                "bwiSyncedAt":   now_ms,
                "products":      [],
                "notes":         "",
                "log":           [{"ts": now_ms, "msg": "Importato da BWI"}],
                "createdAt":     now_ms,
                "updatedAt":     now_ms,
            }
            _update_primary_contact(new_c)
            contacts.append(new_c)
            by_bwi[cid] = new_c
            stats["new"] += 1
            stats["new_list"].append(new_c["company"])

    # Pulisci bwiStatus='updated' dei contatti non più nel XLSX (non li elimina)
    # Solo i 'new' non trovati in XLSX rimangono 'new' (potrebbero essere stati
    # aggiunti manualmente)

    return stats


def _update_primary_contact(c: dict):
    """Popola contactName/contactEmail/contactTitle dal primo contatto."""
    contacts = c.get("contacts", [])
    if contacts:
        first = contacts[0]
        c["contactName"]  = first.get("name","")
        c["contactTitle"] = first.get("title","")
        c["contactEmail"] = first.get("email","")


def send_digest(stats: dict):
    if not BREVO_API_KEY:
        print("  (Brevo API key non configurata — digest non inviato)")
        return
    if stats["new"] == 0 and stats["updated"] == 0:
        print("  (Nessun cambiamento — digest non inviato)")
        return

    new_list_html = "".join(
        f"<li>{n}</li>" for n in stats["new_list"][:30]
    )
    upd_list_html = "".join(
        f"<li>{n}</li>" for n in stats["updated_list"][:20]
    )
    more_new = f"<li>... e altri {len(stats['new_list'])-30}</li>" if len(stats["new_list"]) > 30 else ""
    more_upd = f"<li>... e altri {len(stats['updated_list'])-20}</li>" if len(stats["updated_list"]) > 20 else ""

    today = datetime.now(timezone.utc).strftime("%d/%m/%Y")
    html = f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
  <h2 style="color:#4F1A2A;margin-bottom:4px">BWI Sync — {today}</h2>
  <p style="color:#666;font-size:13px">Riepilogo aggiornamento settimanale importatori</p>
  <hr style="border:none;border-top:1px solid #eee;margin:16px 0">

  <div style="display:flex;gap:16px;margin-bottom:20px">
    <div style="flex:1;background:#e8f5e9;border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:#2e7d32">{stats['new']}</div>
      <div style="font-size:12px;color:#2e7d32;font-weight:600">NUOVI</div>
    </div>
    <div style="flex:1;background:#fff3e0;border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:#e65100">{stats['updated']}</div>
      <div style="font-size:12px;color:#e65100;font-weight:600">AGGIORNATI</div>
    </div>
    <div style="flex:1;background:#f5f5f5;border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:#666">{stats['skipped']}</div>
      <div style="font-size:12px;color:#666;font-weight:600">INVARIATI</div>
    </div>
  </div>

  {'<h3 style="color:#2e7d32;font-size:14px">Nuovi importatori</h3><ul style="font-size:13px;color:#333">'+new_list_html+more_new+'</ul>' if stats['new'] else ''}
  {'<h3 style="color:#e65100;font-size:14px">Importatori aggiornati</h3><ul style="font-size:13px;color:#333">'+upd_list_html+more_upd+'</ul>' if stats['updated'] else ''}

  <hr style="border:none;border-top:1px solid #eee;margin:16px 0">
  <p style="font-size:11px;color:#aaa">Siena Wine CRM — BWI Auto-Sync</p>
</div>"""

    payload = {
        "sender": {"email": SENDER_EMAIL, "name": SENDER_NAME},
        "to":     [{"email": DIGEST_TO_EMAIL, "name": DIGEST_TO_NAME}],
        "subject": f"BWI Sync {today} — {stats['new']} nuovi, {stats['updated']} aggiornati",
        "htmlContent": html,
        "trackClicks": False,
        "trackOpens":  False,
    }
    try:
        r = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            json=payload,
            headers={"api-key": BREVO_API_KEY, "Content-Type": "application/json"},
            timeout=15
        )
        if r.status_code in (200, 201):
            print(f"  Digest inviato a {DIGEST_TO_EMAIL}")
        else:
            print(f"  Digest errore {r.status_code}: {r.text[:100]}")
    except Exception as e:
        print(f"  Digest fallito: {e}")


MANIFEST_FILE = os.path.join(SCRIPT_DIR, "bestwine_output", "last_sync_ids.json")


def main():
    print("=" * 60)
    print("  BWI Sync CRM")
    print("=" * 60)

    # 1. Verifica file
    if not os.path.exists(ALL_XLSX):
        print(f"ERRORE: {ALL_XLSX} non trovato. Esegui prima bestwine_scraper.py + merge_all.py")
        sys.exit(1)
    if not os.path.exists(CRM_FILE):
        print(f"ERRORE: {CRM_FILE} non trovato.")
        sys.exit(1)

    # 2. Carica XLSX — filtra solo i CompId del run corrente se il manifest esiste
    print(f"\n1. Carico {os.path.basename(ALL_XLSX)}...")
    xlsx_companies = load_xlsx_companies(ALL_XLSX)
    print(f"   {len(xlsx_companies):,} aziende totali nel XLSX")

    if os.path.exists(MANIFEST_FILE):
        with open(MANIFEST_FILE) as f:
            recent_ids = set(json.load(f))
        xlsx_companies = {k: v for k, v in xlsx_companies.items() if k in recent_ids}
        print(f"   Filtrato a {len(xlsx_companies):,} aziende da questo run (manifest)")
    else:
        print(f"   Manifest non trovato — sync di tutte le {len(xlsx_companies):,} aziende")
        print(f"   ⚠ Prima esecuzione: potrebbe aggiungere molti contatti al CRM")

    # 3. Carica CRM
    print(f"\n2. Carico crm.json...")
    crm_data = load_crm(CRM_FILE)
    n_before = len(crm_data.get("contacts", []))
    print(f"   {n_before} contatti esistenti nel CRM")

    # 4. Sync
    print(f"\n3. Sync in corso...")
    now_ms = int(time.time() * 1000)
    stats  = sync(crm_data, xlsx_companies, now_ms)

    n_after = len(crm_data.get("contacts", []))
    print(f"   Nuovi:      {stats['new']:>5}")
    print(f"   Aggiornati: {stats['updated']:>5}")
    print(f"   Invariati:  {stats['skipped']:>5}")
    print(f"   Totale CRM: {n_after}")

    # 5. Salva
    print(f"\n4. Salvo crm.json...")
    save_crm(CRM_FILE, crm_data)
    print(f"   Salvato.")

    # 6. Digest
    print(f"\n5. Invio digest...")
    send_digest(stats)

    print(f"\nDONE.")


if __name__ == "__main__":
    main()
