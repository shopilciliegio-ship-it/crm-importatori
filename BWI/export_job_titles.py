"""
export_job_titles.py
====================
Estrae tutti i job title unici dai file Excel in bestwine_output/
e li esporta in un file CSV ordinato per frequenza.

Esegui: python export_job_titles.py

Output: job_titles.csv — aprilo in Excel, mettili in ordine di priorità,
        poi caricalo nel CRM per la logica di selezione destinatario.
"""
import os
import csv
from collections import Counter
from openpyxl import load_workbook

OUTPUT_FOLDER = "bestwine_output"
OUTPUT_CSV    = "job_titles.csv"

# Indice colonna Contact_Title (colonna W, indice 23 da 0)
FIELDNAMES = [
    "CompanyName", "BrandName", "CompId",
    "Country", "City", "State", "StreetAddress", "PostalCode",
    "Website", "Type", "ProdType", "Employee", "Sales",
    "Company_Email", "Phone", "Founded", "RegistrationNumber",
    "Linkedin", "Facebook", "Instagram", "Twitter", "Youtube",
    "Contact_Name", "Contact_Title", "Contact_Email",
    "Contact_Phone", "Contact_Linkedin"
]
TITLE_IDX = FIELDNAMES.index("Contact_Title")  # = 23

files = sorted([
    f for f in os.listdir(OUTPUT_FOLDER)
    if f.endswith(".xlsx")
    and "_OLD" not in f
    and "CORROTTO" not in f
    and f != "bestwine_ALL.xlsx"
])

print(f"📂 Scansione {len(files)} file Excel...\n")

title_counter = Counter()
total_contacts = 0
files_read = 0

for filename in files:
    path = os.path.join(OUTPUT_FOLDER, filename)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = str(row[TITLE_IDX] or "").strip()
            if title and title.lower() not in ("none", "-", ""):
                title_counter[title] += 1
                total_contacts += 1
        wb.close()
        files_read += 1
    except Exception as e:
        print(f"  ⚠️  {filename}: {e}")

print(f"✅ Letti {files_read} file")
print(f"   Contatti con job title: {total_contacts:,}")
print(f"   Job title unici: {len(title_counter):,}")
print()

# Scrivi CSV ordinato per frequenza (più comuni prima)
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f, delimiter=";")
    writer.writerow([
        "Priorità (compila tu)",
        "Job Title",
        "Quante volte appare",
        "% sul totale"
    ])
    for title, count in title_counter.most_common():
        pct = count / total_contacts * 100
        writer.writerow(["", title, count, f"{pct:.1f}%"])

print(f"📄 Esportato: {OUTPUT_CSV}")
print(f"   Apri in Excel, assegna la priorità nella colonna A")
print(f"   (1 = massima priorità, 2 = seconda scelta, ecc.)")
print(f"   Lascia vuota la priorità per i titoli che vuoi ignorare.")
print()

# Anteprima top 30
print("TOP 30 JOB TITLE PIÙ FREQUENTI:")
print(f"{'#':<4} {'Job Title':<45} {'N':>6} {'%':>6}")
print("-" * 65)
for i, (title, count) in enumerate(title_counter.most_common(30), 1):
    pct = count / total_contacts * 100
    print(f"{i:<4} {title:<45} {count:>6} {pct:>5.1f}%")
