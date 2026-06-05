"""
recovery.py
===========
Trova tutte le righe nei file Excel con dati chiave mancanti
(Website O Company_Email vuoti) e le riscarica dall'API BWI.

Esegui NELLA STESSA CARTELLA di bestwine_scraper.py:
    python recovery.py

I dati aggiornati vengono scritti direttamente nei file Excel esistenti.
Le righe aggiornate vengono evidenziate in AZZURRO per distinguerle.
"""

import os
import re
import time
import threading
import requests
from requests.adapters import HTTPAdapter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# COPIA QUI LE STESSE CREDENZIALI DI bestwine_scraper.py
# ============================================================
COOKIE = ""  # incolla qui il cookie di sessione BWI (non committare)

BDN_ACCESS = ""  # incolla qui il BDN access token (non committare)
BDN_ID     = "13229"       # es. 13229
BDN_NAME   = "info@sienawine.it"     # es. info@sienawine.it
# ============================================================

OUTPUT_FOLDER = "bestwine_output"
BASE_URL      = "https://api.bestwineimporters.com/api/v1"

HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Content-Type": "application/json",
    "Cookie": COOKIE,
    "Bdn-Access": BDN_ACCESS,
    "Bdn-Id": BDN_ID,
    "Bdn-Name": BDN_NAME,
    "Origin": "https://app.bestwineimporters.com",
    "Referer": "https://app.bestwineimporters.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Connection": "close"
}

FIELDNAMES = [
    "CompanyName", "BrandName", "CompId",
    "Country", "City", "State", "StreetAddress", "PostalCode",
    "Website", "Type", "ProdType", "Employee", "Sales",
    "Company_Email", "Phone", "Founded", "RegistrationNumber",
    "Linkedin", "Facebook", "Instagram", "Twitter", "Youtube",
    "Contact_Name", "Contact_Title", "Contact_Email",
    "Contact_Phone", "Contact_Linkedin"
]

# Colonne che identificano una riga con profilo NON scaricato
# Questi campi vengono SOLO da get_profile(), mai dalla lista base.
# Se sono tutti vuoti = get_profile() è andato in timeout su questa riga.
PROFILE_COLS = ["Founded", "RegistrationNumber", "Linkedin", "Facebook", "Instagram"]
# Almeno uno di questi deve essere presente per considerare il profilo scaricato
# (alcune aziende non hanno social, ma quasi tutte hanno almeno Founded o RegNum)

# Colore righe recuperate (azzurro)
RECOVER_FILL = PatternFill("solid", fgColor="BDD7EE")
RECOVER_FONT = Font(color="1F4E79")

# Indici colonne per accesso rapido
COL = {name: i for i, name in enumerate(FIELDNAMES)}
COMPID_IDX  = COL["CompId"]
WEBSITE_IDX = COL["Website"]
EMAIL_IDX   = COL["Company_Email"]
PHONE_IDX   = COL["Phone"]


# ── API con hard timeout via thread ──────────────────────────
def make_session():
    s = requests.Session()
    s.mount("https://", HTTPAdapter(max_retries=0))
    s.mount("http://",  HTTPAdapter(max_retries=0))
    return s

def call_api(url, payload, timeout_sec=10):
    result, error = [None], [None]
    def _call():
        try:
            with make_session() as sess:
                r = sess.post(url, json=payload, headers=HEADERS,
                              timeout=(5, timeout_sec))
                result[0] = r
        except Exception as e:
            error[0] = e
    t = threading.Thread(target=_call, daemon=True)
    t.start()
    t.join(timeout=timeout_sec + 3)
    if t.is_alive():
        return None, TimeoutError(f"Hard timeout {timeout_sec+3}s")
    if error[0]:
        return None, error[0]
    return result[0], None

def check_auth():
    r, err = call_api(f"{BASE_URL}/customSearch/", {
        "nocache": False, "startAt": 0, "endAt": 1,
        "sortBy": "", "sortDir": "",
        "filters": {"FilterKeyword":[],"FilterCategories":[],
                    "FilterSubCategories":[],"FilterTypes":[],
                    "FilterOrigin":[],"FilterCountry":[],
                    "FilterContinent":[],"FilterEmployee":[],"FilterSales":[]},
        "id": None, "name": None, "date": None
    }, timeout_sec=8)
    if err: return False, str(err)
    if r.status_code in (401, 403): return False, "Token scaduto (401/403)"
    if r.status_code == 200: return True, "OK"
    return False, f"Status {r.status_code}"

def get_profile(comp_id):
    r, err = call_api(f"{BASE_URL}/profileinfo/",
                      {"compID": comp_id}, timeout_sec=8)
    if err: return {}
    if r.status_code != 200: return {}
    try: return r.json()
    except: return {}

def get_leads(comp_id):
    r, err = call_api(f"{BASE_URL}/leads/",
                      {"compID": comp_id}, timeout_sec=8)
    if err: return []
    if r.status_code != 200: return []
    try:
        data = r.json()
        if isinstance(data, list): return data
        if isinstance(data, dict):
            return data.get("leads", data.get("resListLeads", []))
    except: pass
    return []

def extract_phone(phones):
    if not phones or not isinstance(phones, dict): return ""
    for key in ["fixed_line", "fixed"]:
        nums = phones.get(key, [])
        if nums and isinstance(nums, list): return nums[0]
    return ""

def clean(val):
    s = str(val or "")
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', s)


# ── STEP 1: scansiona tutti i file e trova righe incomplete ──
def scan_incomplete():
    files = sorted([
        f for f in os.listdir(OUTPUT_FOLDER)
        if f.endswith(".xlsx") and "CORROTTO" not in f
        and f != "bestwine_ALL.xlsx"
    ])
    print(f"📂 Scansione {len(files)} file...\n")

    incomplete = []  # lista di (filepath, row_idx, comp_id, company_name)

    for filename in files:
        path = os.path.join(OUTPUT_FOLDER, filename)
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            file_incomplete = 0
            for row_idx, row in enumerate(rows, start=2):
                comp_id  = str(row[COMPID_IDX] or "").strip()
                if not comp_id: continue

                # Considera incompleta se TUTTI i campi profilo sono vuoti
                # (significa che get_profile() non ha mai risposto per questa riga)
                profile_empty = all(
                    not str(row[COL[c]] or "").strip()
                    for c in PROFILE_COLS
                )
                if profile_empty:
                    company = str(row[COL["CompanyName"]] or "")
                    incomplete.append((path, row_idx, comp_id, company))
                    file_incomplete += 1

            if file_incomplete:
                print(f"  ⚠️  {filename}: {file_incomplete} righe incomplete")

        except Exception as e:
            print(f"  ❌ {filename}: {e}")

    return incomplete


# ── STEP 2: riscarica e aggiorna ──────────────────────────────
def recover(incomplete):
    total = len(incomplete)
    print(f"\n🔄 Recovery di {total} righe...\n")

    # Raggruppa per file per aprire/chiudere ogni workbook una volta sola
    by_file = {}
    for (path, row_idx, comp_id, company) in incomplete:
        by_file.setdefault(path, []).append((row_idx, comp_id, company))

    recovered = 0
    failed    = 0
    skipped   = 0

    for path, entries in by_file.items():
        try:
            wb = load_workbook(path)
            ws = wb.active
            modified = False

            for row_idx, comp_id, company in entries:
                print(f"  [{recovered+failed+skipped+1}/{total}] {company} (ID: {comp_id})")

                profile = get_profile(comp_id)
                time.sleep(0.3)

                if not profile:
                    print(f"    ⏱  Profilo non disponibile — salto")
                    skipped += 1
                    continue

                leads = get_leads(comp_id)
                time.sleep(0.3)

                social   = profile.get("social", {})
                phones   = profile.get("phones", {})
                datadnb  = profile.get("datadnb", {})
                location = profile.get("location", {})

                # Dati base aggiornati
                updated = {
                    "BrandName":          profile.get("brandname", ""),
                    "City":               location.get("City", ""),
                    "State":              location.get("State", ""),
                    "StreetAddress":      location.get("Address", ""),
                    "PostalCode":         location.get("Postal", ""),
                    "Website":            profile.get("website", ""),
                    "Employee":           datadnb.get("employees", ""),
                    "Sales":              datadnb.get("sales", ""),
                    "Company_Email":      profile.get("email", ""),
                    "Phone":              extract_phone(phones),
                    "Founded":            datadnb.get("Founded", ""),
                    "RegistrationNumber": datadnb.get("registrationNumber", ""),
                    "Linkedin":           social.get("Linkedin", ""),
                    "Facebook":           social.get("Facebook", ""),
                    "Instagram":          social.get("Instagram", ""),
                    "Twitter":            social.get("Twitter", ""),
                    "Youtube":            social.get("Youtube", ""),
                }

                # Aggiorna la riga esistente nel foglio
                for col_name, value in updated.items():
                    col_idx = COL[col_name] + 1  # openpyxl è 1-based
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = clean(value)

                # Se ci sono leads E la riga non ha già un contatto,
                # aggiorna il primo contatto (Contact_Name, ecc.)
                if leads:
                    lead = leads[0]
                    ws.cell(row=row_idx, column=COL["Contact_Name"]+1).value  = clean(lead.get("FullName",""))
                    ws.cell(row=row_idx, column=COL["Contact_Title"]+1).value = clean(lead.get("Position",""))
                    ws.cell(row=row_idx, column=COL["Contact_Phone"]+1).value = clean(lead.get("Phone",""))
                    ws.cell(row=row_idx, column=COL["Contact_Linkedin"]+1).value = clean(lead.get("LinkedIn",""))

                # Colora la riga in azzurro = recuperata
                for col_idx in range(1, len(FIELDNAMES)+1):
                    ws.cell(row=row_idx, column=col_idx).fill = RECOVER_FILL
                    ws.cell(row=row_idx, column=col_idx).font = RECOVER_FONT

                modified = True
                recovered += 1
                print(f"    ✅ Aggiornato — email:{profile.get('email','?')} web:{profile.get('website','?')}")

            if modified:
                wb.save(path)
                print(f"  💾 Salvato: {os.path.basename(path)}\n")

        except Exception as e:
            print(f"  ❌ Errore su {os.path.basename(path)}: {e}")
            failed += 1

    print(f"""
{'='*50}
✅ Recovery completato!
   Recuperate:  {recovered}
   Non disponibili (skip): {skipped}
   Errori file: {failed}
{'='*50}
Righe recuperate evidenziate in AZZURRO nei file Excel.
Riesegui merge_all.py per aggiornare bestwine_ALL.xlsx.
""")


# ── MAIN ─────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 50)
    print("BWI Recovery — dati mancanti")
    print("=" * 50)

    if not COOKIE or not BDN_ACCESS:
        print("❌ Inserisci COOKIE e BDN_ACCESS in cima al file!")
        exit(1)

    print("\n🔑 Verifica autenticazione...")
    ok, msg = check_auth()
    if not ok:
        print(f"❌ Autenticazione fallita: {msg}")
        print("   Aggiorna COOKIE e BDN_ACCESS e riprova.")
        exit(1)
    print(f"✅ Autenticazione OK\n")

    incomplete = scan_incomplete()

    if not incomplete:
        print("\n✅ Nessuna riga incompleta trovata — tutto a posto!")
        exit(0)

    print(f"\nTotale righe da recuperare: {len(incomplete)}")
    confirm = input("Procedere? [s/N] ").strip().lower()
    if confirm != 's':
        print("Annullato.")
        exit(0)

    recover(incomplete)
